from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import os
from dotenv import load_dotenv
import json
import requests
import hashlib
from typing import Dict

load_dotenv()

app = FastAPI(title="CV Analyzer API")

# Cache để lưu kết quả phân tích
analysis_cache: Dict[str, dict] = {}

# CORS middleware để frontend có thể gọi API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# API keys
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

if not GOOGLE_API_KEY:
    print("WARNING: GOOGLE_API_KEY not found in .env file!")

class CVAnalysisResponse(BaseModel):
    summary: str
    name: str
    email: str
    phone: str
    skills: list[str]
    experience: str
    education: str
    strengths: list[str]
    recommendations: list[str]
    interview_questions: list[str]
    salary_range: str
    career_path: list[str]
    # Job matching fields
    overall_score: int = 0
    skills_score: int = 0
    experience_score: int = 0
    education_score: int = 0
    soft_skills_score: int = 0
    match_percentage: int = 0
    matching_skills: list[str] = []
    missing_skills: list[str] = []
    red_flags: list[str] = []

@app.get("/")
async def root():
    return {
        "message": "CV Analyzer API is running",
        "cached_items": len(analysis_cache)
    }

@app.get("/cache/stats")
async def get_cache_stats():
    """Thống kê cache"""
    return {
        "total_cached": len(analysis_cache),
        "cache_keys": list(analysis_cache.keys())[:10]  # Show first 10
    }

@app.delete("/cache/clear")
async def clear_cache():
    """Xóa toàn bộ cache"""
    global analysis_cache
    count = len(analysis_cache)
    analysis_cache = {}
    return {"message": f"Đã xóa {count} items từ cache"}

@app.post("/batch-analyze")
async def batch_analyze_cvs(
    files: list[UploadFile] = File(...),
    model: str = "gemini-2.0-flash",
    job_description: str = Form("")
):
    """
    Upload nhiều CV cùng lúc và phân tích
    """
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="Tối đa 10 CV mỗi lần")
    
    results = []
    
    for file in files:
        try:
            # Đọc file
            content = await file.read()
            
            # Extract text
            if file.filename.endswith('.txt'):
                cv_text = content.decode('utf-8')
            elif file.filename.endswith('.pdf'):
                cv_text = extract_text_from_pdf(content)
            elif file.filename.endswith('.docx'):
                cv_text = extract_text_from_docx(content)
            else:
                continue
            
            # Phân tích
            analysis = await analyze_cv_with_ai(cv_text, model, job_description)
            
            results.append({
                "filename": file.filename,
                "analysis": analysis.dict()
            })
            
        except Exception as e:
            print(f"Error processing {file.filename}: {e}")
            results.append({
                "filename": file.filename,
                "error": str(e)
            })
    
    return {"results": results, "total": len(results)}

@app.post("/export-excel")
async def export_to_excel(data: dict):
    """
    Export kết quả batch analysis ra Excel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        wb = Workbook()
        ws = wb.active
        ws.title = "CV Analysis Results"
        
        # Header
        headers = [
            "Tên", "Email", "Phone", "Điểm Tổng", "Skills Score", 
            "Experience Score", "Education Score", "Match %", 
            "Skills", "Red Flags", "Salary Range"
        ]
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data
        results = data.get("results", [])
        for row_idx, result in enumerate(results, 2):
            analysis = result.get("analysis", {})
            
            ws.cell(row=row_idx, column=1, value=analysis.get("name", ""))
            ws.cell(row=row_idx, column=2, value=analysis.get("email", ""))
            ws.cell(row=row_idx, column=3, value=analysis.get("phone", ""))
            ws.cell(row=row_idx, column=4, value=analysis.get("overall_score", 0))
            ws.cell(row=row_idx, column=5, value=analysis.get("skills_score", 0))
            ws.cell(row=row_idx, column=6, value=analysis.get("experience_score", 0))
            ws.cell(row=row_idx, column=7, value=analysis.get("education_score", 0))
            ws.cell(row=row_idx, column=8, value=analysis.get("match_percentage", 0))
            ws.cell(row=row_idx, column=9, value=", ".join(analysis.get("skills", [])))
            ws.cell(row=row_idx, column=10, value=", ".join(analysis.get("red_flags", [])))
            ws.cell(row=row_idx, column=11, value=analysis.get("salary_range", ""))
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=cv_analysis_results.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi export Excel: {str(e)}")

@app.get("/models")
async def get_available_models():
    """Lấy danh sách models có sẵn"""
    return {
        "models": [
            {
                "id": "gemini-2.0-flash",
                "name": "Gemini 2.0 Flash",
                "provider": "Google",
                "description": "Nhanh, miễn phí, phù hợp cho hầu hết CV",
                "icon": "⚡"
            },
            {
                "id": "gemini-2.5-flash",
                "name": "Gemini 2.5 Flash",
                "provider": "Google",
                "description": "Phiên bản mới nhất, nhanh và chính xác hơn",
                "icon": "🚀"
            },
            {
                "id": "gemini-2.5-pro",
                "name": "Gemini 2.5 Pro",
                "provider": "Google",
                "description": "Phân tích sâu, chi tiết nhất",
                "icon": "💎"
            },
            {
                "id": "openrouter-claude",
                "name": "Claude 3.5 Sonnet (OpenRouter)",
                "provider": "OpenRouter",
                "description": "Claude qua OpenRouter - Miễn phí với credits",
                "icon": "🔥"
            },
            {
                "id": "openrouter-gpt4",
                "name": "GPT-4 Turbo (OpenRouter)",
                "provider": "OpenRouter",
                "description": "GPT-4 qua OpenRouter - Miễn phí với credits",
                "icon": "🤖"
            },
            {
                "id": "openrouter-llama",
                "name": "Llama 3.1 70B (OpenRouter)",
                "provider": "OpenRouter",
                "description": "Llama 3.1 - Miễn phí, mạnh mẽ",
                "icon": "🦙"
            }
        ]
    }

@app.post("/analyze-cv", response_model=CVAnalysisResponse)
async def analyze_cv(
    file: UploadFile = File(...), 
    model: str = "gemini-2.0-flash",
    job_description: str = Form("")
):
    """
    Upload CV file và phân tích nội dung
    Supported models: gemini-2.0-flash, gemini-2.5-flash, gemini-2.5-pro, claude-sonnet
    """
    # Kiểm tra file type
    if not file.filename.endswith(('.pdf', '.docx', '.txt')):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ file PDF, DOCX, TXT")
    
    try:
        print(f"Processing file: {file.filename} with model: {model}")
        
        # Đọc nội dung file
        content = await file.read()
        print(f"File size: {len(content)} bytes")
        
        # Tạo cache key từ file content + model
        cache_key = hashlib.md5(content + model.encode()).hexdigest()
        
        # Kiểm tra cache
        if cache_key in analysis_cache:
            print(f"✅ Cache hit! Returning cached result for {file.filename}")
            return CVAnalysisResponse(**analysis_cache[cache_key])
        
        print(f"❌ Cache miss. Analyzing with AI...")
        
        # Extract text từ file
        if file.filename.endswith('.txt'):
            cv_text = content.decode('utf-8')
        elif file.filename.endswith('.pdf'):
            cv_text = extract_text_from_pdf(content)
        elif file.filename.endswith('.docx'):
            cv_text = extract_text_from_docx(content)
        
        print(f"Extracted text length: {len(cv_text)} characters")
        print(f"Job Description provided: {bool(job_description)}")
        
        # Gọi AI để phân tích
        analysis = await analyze_cv_with_ai(cv_text, model, job_description)
        
        # Lưu vào cache
        analysis_cache[cache_key] = analysis.dict()
        print(f"💾 Saved to cache. Total cached items: {len(analysis_cache)}")
        
        return analysis
    
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý file: {str(e)}")

async def analyze_cv_with_ai(cv_text: str, model: str = "gemini-2.0-flash", job_description: str = "") -> CVAnalysisResponse:
    """
    Sử dụng AI API để phân tích CV
    Hỗ trợ: Gemini (2.0-flash, 2.5-flash, 2.5-pro) và Claude Sonnet
    """
    # Build prompt based on whether JD is provided
    if job_description:
        prompt = f"""
Bạn là chuyên gia phân tích CV và tư vấn tuyển dụng. Phân tích CV và SO SÁNH với Job Description để trả về JSON:

{{
  "summary": "Tóm tắt về ứng viên",
  "name": "Tên",
  "email": "Email",
  "phone": "Phone",
  "skills": ["skill1", "skill2"],
  "experience": "Kinh nghiệm",
  "education": "Học vấn",
  "strengths": ["điểm mạnh"],
  "recommendations": ["gợi ý"],
  "interview_questions": ["câu hỏi 1", "câu hỏi 2", "..."],
  "salary_range": "Mức lương đề xuất",
  "career_path": ["bước 1", "bước 2"],
  
  "overall_score": 85,
  "skills_score": 90,
  "experience_score": 80,
  "education_score": 85,
  "soft_skills_score": 80,
  "match_percentage": 85,
  "matching_skills": ["skill phù hợp với JD"],
  "missing_skills": ["skill thiếu so với JD"],
  "red_flags": ["Gap 2 năm không làm việc", "Đổi việc 5 lần trong 3 năm", "Thiếu kỹ năng bắt buộc X"]
}}

CHẤM ĐIỂM (0-100):
- overall_score: Tổng điểm
- skills_score: Kỹ năng (30%)
- experience_score: Kinh nghiệm (30%)
- education_score: Học vấn (20%)
- soft_skills_score: Kỹ năng mềm (20%)
- match_percentage: % phù hợp với JD

RED FLAGS (nếu có):
- Gap trong CV
- Job hopping (đổi việc quá nhiều)
- Thiếu yêu cầu bắt buộc
- Overselling

Job Description:
{job_description}

CV Content:
{cv_text}
"""
    else:
        prompt = f"""
Bạn là chuyên gia phân tích CV. Phân tích CV và trả về JSON:

{{
  "summary": "Tóm tắt",
  "name": "Tên",
  "email": "Email",
  "phone": "Phone",
  "skills": ["skill1"],
  "experience": "Kinh nghiệm",
  "education": "Học vấn",
  "strengths": ["điểm mạnh"],
  "recommendations": ["gợi ý"],
  "interview_questions": ["câu hỏi"],
  "salary_range": "Mức lương",
  "career_path": ["bước"],
  "overall_score": 0,
  "skills_score": 0,
  "experience_score": 0,
  "education_score": 0,
  "soft_skills_score": 0,
  "match_percentage": 0,
  "matching_skills": [],
  "missing_skills": [],
  "red_flags": []
}}

CV Content:
{cv_text}
"""
    
    try:
        if model.startswith("gemini"):
            # Gọi Gemini API
            return await analyze_with_gemini(prompt, model, job_description)
        elif model == "claude-sonnet":
            # Gọi Claude API
            return await analyze_with_claude(prompt, job_description)
        elif model.startswith("openrouter"):
            # Gọi OpenRouter API
            return await analyze_with_openrouter(prompt, model, job_description)
        else:
            raise HTTPException(status_code=400, detail=f"Model không hỗ trợ: {model}")
    
    except Exception as e:
        print(f"Full error: {e}")
        raise HTTPException(status_code=500, detail=f"Lỗi phân tích AI: {str(e)}")

async def analyze_with_gemini(prompt: str, model: str, job_description: str = "") -> CVAnalysisResponse:
    """Phân tích CV với Gemini"""
    url = f"https://generativelanguage.googleapis.com/v1/models/{model}:generateContent?key={GOOGLE_API_KEY}"
    
    headers = {"Content-Type": "application/json"}
    
    data = {
        "contents": [{
            "parts": [{"text": prompt}]
        }],
        "generationConfig": {
            "temperature": 0.3,
            "maxOutputTokens": 2048,
        }
    }
    
    response = requests.post(url, headers=headers, json=data)
    response.raise_for_status()
    
    result = response.json()
    response_text = result["candidates"][0]["content"]["parts"][0]["text"].strip()
    
    # Loại bỏ markdown code block
    if response_text.startswith("```json"):
        response_text = response_text[7:]
    if response_text.startswith("```"):
        response_text = response_text[3:]
    if response_text.endswith("```"):
        response_text = response_text[:-3]
    
    response_text = response_text.strip()
    
    # Clean invalid control characters
    import re
    response_text = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', response_text)
    
    try:
        parsed_result = json.loads(response_text)
    except json.JSONDecodeError as e:
        print(f"JSON Parse Error: {e}")
        print(f"Response text: {response_text[:500]}...")
        raise HTTPException(status_code=500, detail=f"Lỗi parse JSON từ AI: {str(e)}")
    
    # Normalize data - convert arrays to strings if needed
    if isinstance(parsed_result.get('experience'), list):
        parsed_result['experience'] = '\n\n'.join([
            f"{exp.get('title', '')} - {exp.get('company', '')}\n{exp.get('duration', '')}\n{exp.get('description', '')}"
            for exp in parsed_result['experience']
        ])
    
    if isinstance(parsed_result.get('education'), list):
        parsed_result['education'] = '\n'.join([
            f"{edu.get('degree', '')} - {edu.get('major', '')}\n{edu.get('school', '')} ({edu.get('year', '')})"
            for edu in parsed_result['education']
        ])
    
    return CVAnalysisResponse(**parsed_result)

async def analyze_with_claude(prompt: str, job_description: str = "") -> CVAnalysisResponse:
    """Phân tích CV với Claude Sonnet"""
    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
    
    if not ANTHROPIC_API_KEY or ANTHROPIC_API_KEY == "your_anthropic_api_key_here":
        raise HTTPException(
            status_code=400, 
            detail="Claude API key chưa được cấu hình. Vui lòng dùng Gemini models hoặc thêm ANTHROPIC_API_KEY vào file .env"
        )
    
    url = "https://api.anthropic.com/v1/messages"
    
    headers = {
        "Content-Type": "application/json",
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01"
    }
    
    data = {
        "model": "claude-3-5-sonnet-20241022",
        "max_tokens": 2048,
        "temperature": 0.3,
        "messages": [
            {"role": "user", "content": prompt}
        ]
    }
    
    try:
        response = requests.post(url, headers=headers, json=data, timeout=60)
        
        if response.status_code != 200:
            error_detail = response.json() if response.text else {}
            print(f"Claude API Error: {response.status_code} - {error_detail}")
            raise HTTPException(
                status_code=400,
                detail=f"Claude API error: {error_detail.get('error', {}).get('message', 'Invalid API key or request')}"
            )
        
        result = response.json()
        response_text = result["content"][0]["text"].strip()
        
        # Loại bỏ markdown code block
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        response_text = response_text.strip()
        
        # Clean invalid control characters
        import re
        response_text = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', response_text)
        
        try:
            parsed_result = json.loads(response_text)
        except json.JSONDecodeError as e:
            print(f"JSON Parse Error: {e}")
            print(f"Response text: {response_text[:500]}...")
            raise HTTPException(status_code=500, detail=f"Lỗi parse JSON từ AI: {str(e)}")
        
        # Normalize data - convert arrays to strings if needed
        if isinstance(parsed_result.get('experience'), list):
            parsed_result['experience'] = '\n\n'.join([
                f"{exp.get('title', '')} - {exp.get('company', '')}\n{exp.get('duration', '')}\n{exp.get('description', '')}"
                for exp in parsed_result['experience']
            ])
        
        if isinstance(parsed_result.get('education'), list):
            parsed_result['education'] = '\n'.join([
                f"{edu.get('degree', '')} - {edu.get('major', '')}\n{edu.get('school', '')} ({edu.get('year', '')})"
                for edu in parsed_result['education']
            ])
        
        return CVAnalysisResponse(**parsed_result)
    
    except requests.exceptions.RequestException as e:
        print(f"Request error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Lỗi kết nối Claude API: {str(e)}"
        )

async def analyze_with_openrouter(prompt: str, model: str, job_description: str = "") -> CVAnalysisResponse:
    """Phân tích CV với OpenRouter (hỗ trợ nhiều models)"""
    if not OPENROUTER_API_KEY or OPENROUTER_API_KEY == "your_openrouter_api_key_here":
        raise HTTPException(
            status_code=400,
            detail="OpenRouter API key chưa được cấu hình. Lấy miễn phí tại: https://openrouter.ai/keys"
        )
    
    # Map model IDs
    model_mapping = {
        "openrouter-claude": "anthropic/claude-3.5-sonnet",
        "openrouter-gpt4": "openai/gpt-4-turbo",
        "openrouter-llama": "meta-llama/llama-3.1-70b-instruct"
    }
    
    openrouter_model = model_mapping.get(model, "anthropic/claude-3.5-sonnet")
    
    url = "https://openrouter.ai/api/v1/chat/completions"
    
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
        "HTTP-Referer": "http://localhost:3000",
        "X-Title": "CV Analyzer"
    }
    
    data = {
        "model": openrouter_model,
        "messages": [
            {
                "role": "user",
                "content": prompt
            }
        ],
        "temperature": 0.3,
        "max_tokens": 2048
    }
    
    try:
        response = requests.post(url, headers=headers, json=data, timeout=60)
        
        if response.status_code != 200:
            error_detail = response.json() if response.text else {}
            print(f"OpenRouter API Error: {response.status_code} - {error_detail}")
            raise HTTPException(
                status_code=400,
                detail=f"OpenRouter API error: {error_detail.get('error', {}).get('message', 'Invalid request')}"
            )
        
        result = response.json()
        response_text = result["choices"][0]["message"]["content"].strip()
        
        # Loại bỏ markdown code block
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        response_text = response_text.strip()
        
        # Clean invalid control characters
        import re
        response_text = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', response_text)
        
        try:
            parsed_result = json.loads(response_text)
        except json.JSONDecodeError as e:
            print(f"JSON Parse Error: {e}")
            print(f"Response text: {response_text[:500]}...")
            raise HTTPException(status_code=500, detail=f"Lỗi parse JSON từ AI: {str(e)}")
        
        # Normalize data
        if isinstance(parsed_result.get('experience'), list):
            parsed_result['experience'] = '\n\n'.join([
                f"{exp.get('title', '')} - {exp.get('company', '')}\n{exp.get('duration', '')}\n{exp.get('description', '')}"
                for exp in parsed_result['experience']
            ])
        
        if isinstance(parsed_result.get('education'), list):
            parsed_result['education'] = '\n'.join([
                f"{edu.get('degree', '')} - {edu.get('major', '')}\n{edu.get('school', '')} ({edu.get('year', '')})"
                for edu in parsed_result['education']
            ])
        
        return CVAnalysisResponse(**parsed_result)
    
    except requests.exceptions.RequestException as e:
        print(f"Request error: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Lỗi kết nối OpenRouter API: {str(e)}"
        )

def extract_text_from_pdf(content: bytes) -> str:
    """Extract text từ PDF"""
    import io
    from PyPDF2 import PdfReader
    
    pdf_file = io.BytesIO(content)
    reader = PdfReader(pdf_file)
    
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    
    return text

def extract_text_from_docx(content: bytes) -> str:
    """Extract text từ DOCX"""
    import io
    from docx import Document
    
    docx_file = io.BytesIO(content)
    doc = Document(docx_file)
    
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    
    return text

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
